import streamlit as st
import pandas as pd
import plotly.express as px
import warnings
import re
import hashlib
from datetime import datetime
import os
from pathlib import Path
import glob
from sqlalchemy import create_engine, text
import comparison as cmp
from comparison_excel import build_comparison_workbook

# --- CONFIGURATION ---
st.set_page_config(page_title="Playbet Performance", layout="wide")
warnings.filterwarnings('ignore')

# Startup checkpoint — proves the app booted past imports. If the screen stays
# blank/spinning, the hang is in imports; if you see this, the hang is later.
_boot = st.empty()
_boot.info("Starting up…")

BRANCHES = ["Malvern", "Potchefstroom", "Pretoria", "White River", "Randburg"]
month_order = ["January", "February", "March", "April", "May", "June",
               "July", "August", "September", "October", "November", "December"]
def is_blocked(month, year):
    """No months are blocked."""
    return False
YEAR_COLORS = {"2024": "#3498db", "2025": "#e67e22", "2026": "#9b59b6"}
DEPOSIT_YEAR_COLORS = {"2024": "#27ae60", "2025": "#f1c40f", "2026": "#8e44ad"}
GAME_PALETTE = px.colors.qualitative.Vivid

# Define folders
HISTORICAL_FOLDER = "historical_data"
UPLOAD_FOLDER = "uploads"
Path(HISTORICAL_FOLDER).mkdir(exist_ok=True)
Path(UPLOAD_FOLDER).mkdir(exist_ok=True)

# The strict schema every file must conform to before merging
TARGET_COLS = ['Shop', 'Game', 'Deposits', 'GGR', 'Paid Out Sum', 'GW Margin %', 'Net Win', 'Net Win Margin', 'Year', 'Month', 'MonthNum']

# =====================================================================
# NEON PERSISTENCE — manual entries and uploaded CSV rows are stored in
# Neon so they survive Streamlit restarts (session state and the local
# uploads folder are both ephemeral and get wiped on restart).
# =====================================================================
def _get_db_url():
    url = ""
    try:
        url = st.secrets.get("DATABASE_URL", "")
    except Exception:
        url = ""
    if not url:
        url = os.environ.get("DATABASE_URL", "")
    if url.startswith("postgres://"):
        url = url.replace("postgres://", "postgresql://", 1)
    if url.startswith("postgresql+psycopg://"):
        url = url.replace("postgresql+psycopg://", "postgresql://", 1)
    # channel_binding=require makes the psycopg2 handshake hang on some hosts
    # (Streamlit Cloud). sslmode=require alone is sufficient for Neon, so strip it.
    url = url.replace("&channel_binding=require", "").replace("?channel_binding=require&", "?")
    url = url.replace("channel_binding=require", "")
    url = url.replace("?&", "?").rstrip("&")
    return url


DATABASE_URL = _get_db_url()
_engine = None
if DATABASE_URL:
    try:
        _engine = create_engine(
            DATABASE_URL, pool_pre_ping=True,
            connect_args={"connect_timeout": 10},  # fail fast instead of hanging
        )
        with _engine.begin() as _c:
            _c.execute(text("""
                CREATE TABLE IF NOT EXISTS dashboard_manual_entries (
                    id SERIAL PRIMARY KEY,
                    shop VARCHAR(120), game VARCHAR(200),
                    deposits NUMERIC, ggr NUMERIC,
                    year VARCHAR(8), month VARCHAR(20), monthnum INTEGER
                )
            """))
            _c.execute(text("""
                CREATE TABLE IF NOT EXISTS dashboard_uploaded_rows (
                    id SERIAL PRIMARY KEY,
                    source_file VARCHAR(300),
                    shop VARCHAR(120), game VARCHAR(200),
                    deposits NUMERIC, paid_out_sum NUMERIC, ggr NUMERIC,
                    gw_margin NUMERIC, net_win NUMERIC, net_win_margin NUMERIC,
                    year VARCHAR(8), month VARCHAR(20), monthnum INTEGER
                )
            """))
            _c.execute(text("""
                CREATE TABLE IF NOT EXISTS dashboard_upload_hashes (
                    content_hash VARCHAR(64) PRIMARY KEY,
                    source_file VARCHAR(300),
                    uploaded_at TIMESTAMP DEFAULT NOW()
                )
            """))
            # Full slip-summary rows (per cashier) — kept, NOT collapsed — so the
            # comparison section can do betslip/cashier head-to-head and GWM%.
            _c.execute(text("""
                CREATE TABLE IF NOT EXISTS dashboard_slip_rows (
                    id SERIAL PRIMARY KEY,
                    source_file VARCHAR(300),
                    shop VARCHAR(120), cashier VARCHAR(200),
                    bet_slips INTEGER, paid_in NUMERIC, paid_out NUMERIC,
                    gw_margin NUMERIC, net_win NUMERIC,
                    year VARCHAR(8), month VARCHAR(20), monthnum INTEGER
                )
            """))
            # Per game-per-shop rows from the Cash Operations export — kept so
            # the comparison can do game betslip counts and per-game GWM%.
            _c.execute(text("""
                CREATE TABLE IF NOT EXISTS dashboard_game_rows (
                    id SERIAL PRIMARY KEY,
                    source_file VARCHAR(300),
                    shop VARCHAR(120), game VARCHAR(200),
                    paid_in_count INTEGER, paid_in_sum NUMERIC,
                    paid_out_count INTEGER, paid_out_sum NUMERIC, gross_win NUMERIC,
                    year VARCHAR(8), month VARCHAR(20), monthnum INTEGER
                )
            """))
            # (No month/year cleanup — all data is retained.)
    except Exception as e:
        st.sidebar.warning(f"DB connection issue: {e}")

_boot.info("Database ready — loading data…")


def load_manual_entries_from_neon():
    if _engine is None:
        return pd.DataFrame(columns=TARGET_COLS)
    try:
        with _engine.connect() as c:
            rows = c.execute(text(
                "SELECT shop, game, deposits, ggr, year, month, monthnum FROM dashboard_manual_entries"
            )).fetchall()
        if not rows:
            return pd.DataFrame(columns=TARGET_COLS)
        return pd.DataFrame([{
            'Shop': r[0], 'Game': r[1], 'Deposits': float(r[2] or 0), 'Paid Out Sum': 0.0,
            'GGR': float(r[3] or 0), 'GW Margin %': 0.0, 'Net Win': 0.0, 'Net Win Margin': 0.0,
            'Year': r[4], 'Month': r[5], 'MonthNum': int(r[6] or 0)
        } for r in rows])
    except Exception as e:
        st.sidebar.warning(f"Could not load manual entries: {e}")
        return pd.DataFrame(columns=TARGET_COLS)


def save_manual_entry_to_neon(shop, game, deposits, ggr, year, month, monthnum):
    if _engine is None:
        return False
    try:
        with _engine.begin() as c:
            c.execute(text("""
                INSERT INTO dashboard_manual_entries (shop, game, deposits, ggr, year, month, monthnum)
                VALUES (:s, :g, :d, :r, :y, :m, :mn)
            """), {"s": shop, "g": game, "d": float(deposits), "r": float(ggr),
                   "y": year, "m": month, "mn": int(monthnum)})
        return True
    except Exception as e:
        st.sidebar.error(f"Could not save entry: {e}")
        return False


def clear_manual_entries_neon():
    if _engine is None:
        return
    try:
        with _engine.begin() as c:
            c.execute(text("DELETE FROM dashboard_manual_entries"))
    except Exception:
        pass


def _content_hash(df_clean):
    """Stable fingerprint of an upload's meaningful content, order-independent."""
    sig_cols = ['Shop', 'Game', 'Deposits', 'GGR', 'Year', 'Month', 'MonthNum']
    try:
        sig_df = df_clean[sig_cols].copy()
        sig_rows = sorted(
            "|".join(str(v) for v in row)
            for row in sig_df.itertuples(index=False, name=None)
        )
        return hashlib.md5("\n".join(sig_rows).encode("utf-8")).hexdigest()
    except Exception:
        return None


def save_uploaded_rows_to_neon(df_clean, source_file):
    """Save uploaded rows to Neon, refusing duplicates by content.
    Returns one of: 'saved', 'duplicate', 'error'."""
    if _engine is None or df_clean is None or df_clean.empty:
        return "error"

    # Drop only blocked month+year combos (2026 May/June/July) before saving
    df_clean = df_clean[~df_clean.apply(
        lambda r: is_blocked(r['Month'], r['Year']), axis=1)]
    if df_clean.empty:
        return "error"

    content_hash = _content_hash(df_clean)

    # The GGR/Deposits summary must hold ONE set of rows per Month+Year. Two
    # different files for the same month (e.g. a slip export and a cash-ops export)
    # must not both be summed in. So we key on the (Year, Month) combos in this file
    # and replace any existing summary rows for those months before inserting.
    ym_pairs = set(zip(df_clean['Year'].astype(str), df_clean['Month'].astype(str)))

    try:
        with _engine.begin() as c:
            # Duplicate check by content fingerprint (catches renamed re-uploads)
            if content_hash is not None:
                existing = c.execute(
                    text("SELECT source_file FROM dashboard_upload_hashes WHERE content_hash = :h"),
                    {"h": content_hash}
                ).fetchone()
                if existing:
                    return "duplicate"

            # Replace any prior summary rows for the SAME Month+Year (from any file),
            # so a month is only ever counted once in the GGR/Deposits chart.
            for (yy, mm) in ym_pairs:
                c.execute(text(
                    "DELETE FROM dashboard_uploaded_rows WHERE year = :y AND month = :m"),
                    {"y": yy, "m": mm})
            for _, row in df_clean.iterrows():
                c.execute(text("""
                    INSERT INTO dashboard_uploaded_rows
                        (source_file, shop, game, deposits, paid_out_sum, ggr,
                         gw_margin, net_win, net_win_margin, year, month, monthnum)
                    VALUES (:f, :shop, :game, :dep, :po, :ggr, :gw, :nw, :nwm, :y, :m, :mn)
                """), {
                    "f": source_file, "shop": row['Shop'], "game": row['Game'],
                    "dep": float(row['Deposits']), "po": float(row['Paid Out Sum']),
                    "ggr": float(row['GGR']), "gw": float(row['GW Margin %']),
                    "nw": float(row['Net Win']), "nwm": float(row['Net Win Margin']),
                    "y": str(row['Year']), "m": str(row['Month']), "mn": int(row['MonthNum'])
                })

            if content_hash is not None:
                c.execute(text("""
                    INSERT INTO dashboard_upload_hashes (content_hash, source_file)
                    VALUES (:h, :f)
                    ON CONFLICT (content_hash) DO NOTHING
                """), {"h": content_hash, "f": source_file})
        return "saved"
    except Exception as e:
        st.sidebar.warning(f"Could not save uploaded rows: {e}")
        return "error"


def load_uploaded_rows_from_neon():
    if _engine is None:
        return pd.DataFrame(columns=TARGET_COLS)
    try:
        with _engine.connect() as c:
            rows = c.execute(text("""
                SELECT shop, game, deposits, paid_out_sum, ggr, gw_margin,
                       net_win, net_win_margin, year, month, monthnum
                FROM dashboard_uploaded_rows
            """)).fetchall()
        if not rows:
            return pd.DataFrame(columns=TARGET_COLS)
        return pd.DataFrame([{
            'Shop': r[0], 'Game': r[1], 'Deposits': float(r[2] or 0), 'Paid Out Sum': float(r[3] or 0),
            'GGR': float(r[4] or 0), 'GW Margin %': float(r[5] or 0), 'Net Win': float(r[6] or 0),
            'Net Win Margin': float(r[7] or 0), 'Year': r[8], 'Month': r[9], 'MonthNum': int(r[10] or 0)
        } for r in rows])
    except Exception:
        return pd.DataFrame(columns=TARGET_COLS)


# =====================================================================
# FULL-DETAIL INGESTION FOR THE COMPARISON SECTION
# Slip Summary -> per-cashier rows (betslips, paid in/out, GW margin, net win)
# Cash Operations -> per game-per-shop rows (paid-in count = betslip proxy, gross win)
# Both keep every row, tagged with Year/Month from the filename.
# =====================================================================
def _col(df, *names):
    for n in names:
        for c in df.columns:
            if str(c).lower().strip() == n:
                return c
    return None


def is_cash_ops_csv(df):
    """Cash Operations export: has Cashier + Shop + Game + Paid In Count + Gross Win."""
    cols = [str(c).lower().strip() for c in df.columns]
    return ('cashier' in cols and 'shop' in cols and 'game' in cols
            and 'paid in count' in cols and 'gross win' in cols)


def is_raw_betslip_csv(df):
    """Raw per-slip transaction dump (one row per betslip): 'Slip #' + 'Paid In Date'
    + 'Pay-in Shop'. These are huge (100k-300k rows) and must NOT be parsed by the
    folder loaders — they'd exhaust memory and hang the app. They are not a summary."""
    cols = [str(c).lower().strip() for c in df.columns]
    return ('slip #' in cols and 'pay-in shop' in cols and 'paid in date' in cols)


def save_slip_rows_to_neon(df, file_date, source_file):
    """Store every cashier row from a Slip Summary CSV. Returns 'saved'|'duplicate'|'error'."""
    if _engine is None or df is None or df.empty:
        return "error"
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    shop_c = _col(df, 'shop')
    user_c = _col(df, 'user')
    slips_c = _col(df, 'bet slips', 'betslips')
    pin_c = _col(df, 'paid in', 'paidin', 'paid in sum')
    pout_c = _col(df, 'paid out', 'paid out sum')
    gw_c = _col(df, 'gw margin %', 'gw margin')
    nw_c = _col(df, 'net win')
    if not (shop_c and user_c and slips_c):
        return "error"

    year = str(file_date.year)
    month = file_date.strftime('%B')
    monthnum = file_date.month

    rows = []
    for _, r in df.iterrows():
        shop = str(r[shop_c]).strip().replace('Potch', 'Potchefstroom')
        if shop not in BRANCHES:
            continue
        rows.append({
            "shop": shop, "cashier": str(r[user_c]).strip(),
            "slips": int(clean_currency_string(r[slips_c])),
            "pin": clean_currency_string(r[pin_c]) if pin_c else 0.0,
            "pout": clean_currency_string(r[pout_c]) if pout_c else 0.0,
            "gw": clean_currency_string(r[gw_c]) if gw_c else 0.0,
            "nw": clean_currency_string(r[nw_c]) if nw_c else 0.0,
        })
    if not rows:
        return "error"

    # content fingerprint for dedup
    sig = hashlib.md5(("SLIP|" + year + month + "|" + "|".join(
        sorted(f"{x['shop']}:{x['cashier']}:{x['slips']}" for x in rows)
    )).encode()).hexdigest()

    try:
        with _engine.begin() as c:
            exists = c.execute(text(
                "SELECT 1 FROM dashboard_upload_hashes WHERE content_hash = :h"), {"h": sig}).fetchone()
            if exists:
                return "duplicate"
            c.execute(text("DELETE FROM dashboard_slip_rows WHERE year=:y AND month=:m"),
                      {"y": year, "m": month})
            for x in rows:
                c.execute(text("""
                    INSERT INTO dashboard_slip_rows
                      (source_file, shop, cashier, bet_slips, paid_in, paid_out,
                       gw_margin, net_win, year, month, monthnum)
                    VALUES (:f,:shop,:cash,:slips,:pin,:pout,:gw,:nw,:y,:m,:mn)
                """), {"f": source_file, "shop": x["shop"], "cash": x["cashier"],
                       "slips": x["slips"], "pin": x["pin"], "pout": x["pout"],
                       "gw": x["gw"], "nw": x["nw"], "y": year, "m": month, "mn": monthnum})
            c.execute(text("""INSERT INTO dashboard_upload_hashes (content_hash, source_file)
                              VALUES (:h,:f) ON CONFLICT (content_hash) DO NOTHING"""),
                      {"h": sig, "f": source_file})
        return "saved"
    except Exception as e:
        st.sidebar.warning(f"Could not save slip rows: {e}")
        return "error"


def save_game_rows_to_neon(df, file_date, source_file):
    """Store every game/shop row from a Cash Operations CSV. Returns 'saved'|'duplicate'|'error'."""
    if _engine is None or df is None or df.empty:
        return "error"
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    shop_c = _col(df, 'shop')
    game_c = _col(df, 'game')
    pinc_c = _col(df, 'paid in count')
    pins_c = _col(df, 'paid in sum')
    poutc_c = _col(df, 'paid out - count', 'paid out count')
    pouts_c = _col(df, 'paid out sum')
    gw_c = _col(df, 'gross win')
    if not (shop_c and game_c and pinc_c and gw_c):
        return "error"

    year = str(file_date.year)
    month = file_date.strftime('%B')
    monthnum = file_date.month

    rows = []
    for _, r in df.iterrows():
        shop = str(r[shop_c]).strip().replace('Potch', 'Potchefstroom')
        if shop not in BRANCHES:
            continue
        rows.append({
            "shop": shop, "game": clean_game_name(r[game_c]),
            "pinc": int(clean_currency_string(r[pinc_c])),
            "pins": clean_currency_string(r[pins_c]) if pins_c else 0.0,
            "poutc": int(clean_currency_string(r[poutc_c])) if poutc_c else 0,
            "pouts": clean_currency_string(r[pouts_c]) if pouts_c else 0.0,
            "gw": clean_currency_string(r[gw_c]),
        })
    if not rows:
        return "error"

    sig = hashlib.md5(("GAME|" + year + month + "|" + "|".join(
        sorted(f"{x['shop']}:{x['game']}:{x['pinc']}" for x in rows)
    )).encode()).hexdigest()

    try:
        with _engine.begin() as c:
            exists = c.execute(text(
                "SELECT 1 FROM dashboard_upload_hashes WHERE content_hash = :h"), {"h": sig}).fetchone()
            if exists:
                return "duplicate"
            c.execute(text("DELETE FROM dashboard_game_rows WHERE year=:y AND month=:m"),
                      {"y": year, "m": month})
            for x in rows:
                c.execute(text("""
                    INSERT INTO dashboard_game_rows
                      (source_file, shop, game, paid_in_count, paid_in_sum,
                       paid_out_count, paid_out_sum, gross_win, year, month, monthnum)
                    VALUES (:f,:shop,:game,:pinc,:pins,:poutc,:pouts,:gw,:y,:m,:mn)
                """), {"f": source_file, "shop": x["shop"], "game": x["game"],
                       "pinc": x["pinc"], "pins": x["pins"], "poutc": x["poutc"],
                       "pouts": x["pouts"], "gw": x["gw"], "y": year, "m": month, "mn": monthnum})
            c.execute(text("""INSERT INTO dashboard_upload_hashes (content_hash, source_file)
                              VALUES (:h,:f) ON CONFLICT (content_hash) DO NOTHING"""),
                      {"h": sig, "f": source_file})
        return "saved"
    except Exception as e:
        st.sidebar.warning(f"Could not save game rows: {e}")
        return "error"


def load_slip_rows_from_neon():
    if _engine is None:
        return pd.DataFrame()
    try:
        with _engine.connect() as c:
            rows = c.execute(text("""
                SELECT shop, cashier, bet_slips, paid_in, paid_out, gw_margin,
                       net_win, year, month, monthnum FROM dashboard_slip_rows
            """)).fetchall()
        if not rows:
            return pd.DataFrame()
        return pd.DataFrame([{
            'Shop': r[0], 'Cashier': r[1], 'BetSlips': int(r[2] or 0),
            'PaidIn': float(r[3] or 0), 'PaidOut': float(r[4] or 0),
            'GWMargin': float(r[5] or 0), 'NetWin': float(r[6] or 0),
            'Year': r[7], 'Month': r[8], 'MonthNum': int(r[9] or 0)
        } for r in rows])
    except Exception:
        return pd.DataFrame()


def load_game_rows_from_neon():
    if _engine is None:
        return pd.DataFrame()
    try:
        with _engine.connect() as c:
            rows = c.execute(text("""
                SELECT shop, game, paid_in_count, paid_in_sum, paid_out_count,
                       paid_out_sum, gross_win, year, month, monthnum FROM dashboard_game_rows
            """)).fetchall()
        if not rows:
            return pd.DataFrame()
        return pd.DataFrame([{
            'Shop': r[0], 'Game': r[1], 'BetSlips': int(r[2] or 0),
            'PaidIn': float(r[3] or 0), 'PaidOutCount': int(r[4] or 0),
            'PaidOut': float(r[5] or 0), 'GrossWin': float(r[6] or 0),
            'Year': r[7], 'Month': r[8], 'MonthNum': int(r[9] or 0)
        } for r in rows])
    except Exception:
        return pd.DataFrame()


@st.cache_data
def load_comparison_detail_from_folder():
    """Scan the uploads/ folder and build slip + game detail frames for the
    comparison, so months already committed to the repo appear without re-upload.
    Returns (slip_df, game_df)."""
    slip_parts, game_parts = [], []
    for file_path in glob.glob(os.path.join(UPLOAD_FOLDER, "*.csv")):
        try:
            filename = os.path.basename(file_path)
            file_date = extract_date_from_filename(filename)
            if file_date is None:
                continue
            # Peek at the header only — skip giant raw per-slip dumps before loading them.
            head = pd.read_csv(file_path, nrows=0)
            head.columns = [str(c).strip() for c in head.columns]
            if is_raw_betslip_csv(head):
                continue
            df = pd.read_csv(file_path)
            df.columns = [str(c).strip() for c in df.columns]
            year = str(file_date.year); month = file_date.strftime('%B'); mn = file_date.month

            if is_cash_ops_csv(df):
                sub = df.copy()
                sub['Shop'] = sub[_col(sub, 'shop')].astype(str).str.strip().replace({'Potch': 'Potchefstroom'})
                sub = sub[sub['Shop'].isin(BRANCHES)]
                if sub.empty:
                    continue
                betslip_col = _col(sub, 'paid in count')
                paidin_col = _col(sub, 'paid in sum')
                paidout_col = _col(sub, 'paid out sum')
                grosswin_col = _col(sub, 'gross win')
                cashier_col = _col(sub, 'cashier')

                sub['_betslips'] = sub[betslip_col].apply(clean_currency_string)
                sub['_paidin'] = sub[paidin_col].apply(clean_currency_string) if paidin_col else 0.0
                sub['_paidout'] = sub[paidout_col].apply(clean_currency_string) if paidout_col else 0.0
                sub['_grosswin'] = sub[grosswin_col].apply(clean_currency_string)

                # Game detail: one row per game (summed across cashiers)
                game_parts.append(pd.DataFrame({
                    'Shop': sub['Shop'], 'Game': sub[_col(sub, 'game')].apply(clean_game_name),
                    'BetSlips': sub['_betslips'].astype(int),
                    'PaidIn': sub['_paidin'], 'PaidOutCount': 0,
                    'PaidOut': sub['_paidout'], 'GrossWin': sub['_grosswin'],
                    'Year': year, 'Month': month, 'MonthNum': mn}))

                # Cashier detail: derived from the same file (summed across games per cashier),
                # so a single Cash Operations file fills BOTH the games and cashiers sections.
                if cashier_col:
                    cg = sub.groupby([cashier_col, 'Shop'], as_index=False).agg(
                        BetSlips=('_betslips', 'sum'), PaidIn=('_paidin', 'sum'),
                        PaidOut=('_paidout', 'sum'), GrossWin=('_grosswin', 'sum'))
                    cg['GWMargin'] = (cg['GrossWin'] / cg['PaidIn'].replace(0, float('nan')) * 100).round(2).fillna(0.0)
                    slip_parts.append(pd.DataFrame({
                        'Shop': cg['Shop'], 'Cashier': cg[cashier_col].astype(str).str.strip(),
                        'BetSlips': cg['BetSlips'].astype(int),
                        'PaidIn': cg['PaidIn'], 'PaidOut': cg['PaidOut'],
                        'GWMargin': cg['GWMargin'], 'NetWin': cg['GrossWin'],
                        'Year': year, 'Month': month, 'MonthNum': mn}))
            elif is_per_user_csv(df):
                sub = df.copy()
                sub['Shop'] = sub[_col(sub, 'shop')].astype(str).str.strip().replace({'Potch': 'Potchefstroom'})
                sub = sub[sub['Shop'].isin(BRANCHES)]
                if sub.empty:
                    continue
                pin = _col(sub, 'paid in', 'paidin', 'paid in sum')
                pout = _col(sub, 'paid out')
                gw = _col(sub, 'gw margin %', 'gw margin')
                nw = _col(sub, 'net win')
                slip_parts.append(pd.DataFrame({
                    'Shop': sub['Shop'], 'Cashier': sub[_col(sub, 'user')].astype(str).str.strip(),
                    'BetSlips': sub[_col(sub, 'bet slips')].apply(clean_currency_string).astype(int),
                    'PaidIn': sub[pin].apply(clean_currency_string) if pin else 0.0,
                    'PaidOut': sub[pout].apply(clean_currency_string) if pout else 0.0,
                    'GWMargin': sub[gw].apply(clean_currency_string) if gw else 0.0,
                    'NetWin': sub[nw].apply(clean_currency_string) if nw else 0.0,
                    'Year': year, 'Month': month, 'MonthNum': mn}))
        except Exception:
            continue
    slip_df = pd.concat(slip_parts, ignore_index=True) if slip_parts else pd.DataFrame()
    game_df = pd.concat(game_parts, ignore_index=True) if game_parts else pd.DataFrame()
    return slip_df, game_df


@st.cache_data
def load_comparison_detail_from_historical():
    """Parse the 'Games and Users' sheets in the historical_data/ Excel files into
    per-cashier×game rows, so the comparison spans 2024 and 2025 too. These sheets
    carry Game, Shop, User, Bet Slips, Paid In, Paid Out, Gross Win, GW Margin %,
    with 2024 and 2025 rows separated by the First Slip Issued date.
    Returns (slip_df, game_df) in the same shape as the folder loader."""
    import openpyxl
    rows = []

    def _yr(d):
        if hasattr(d, "year"):
            return d.year
        # dates are DD/MM/YY — the YEAR is the LAST two digits, not the first.
        m = re.match(r"\s*\d{1,2}/\d{1,2}/(\d{2})\b", str(d))
        return 2000 + int(m.group(1)) if m else None

    def _month_from_sheet(name):
        for i, m in enumerate(month_order, 1):
            if m.lower()[:3] in name.lower():
                return m, i
        return None, None

    files = glob.glob(os.path.join(HISTORICAL_FOLDER, "*.xlsx")) + \
            glob.glob(os.path.join(HISTORICAL_FOLDER, "*.xls"))
    for fp in files:
        try:
            wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
        except Exception:
            continue
        for sh in wb.sheetnames:
            month, mn = _month_from_sheet(sh)
            if not month:
                continue
            ws = wb[sh]
            it = ws.iter_rows(values_only=True)
            try:
                header = next(it)
            except StopIteration:
                continue
            idx = {str(h).strip().lower(): i for i, h in enumerate(header) if h}
            # Detect the per-cashier×game layout by its COLUMNS, not the sheet name.
            # Some months (e.g. April) sit on a sheet named "Excl users" but still
            # carry the full Game/Shop/User/Bet Slips layout — those must be read.
            if not ("game" in idx and "user" in idx and "shop" in idx and "bet slips" in idx):
                continue

            def g(row, key):
                return row[idx[key]] if key in idx and idx[key] < len(row) else None

            for row in it:
                if not row or row[0] is None:
                    continue
                shop = str(g(row, "shop")).strip().replace("Potch", "Potchefstroom")
                if shop not in BRANCHES:
                    continue
                y = _yr(g(row, "first slip issued"))
                if y not in (2024, 2025):
                    continue
                rows.append({
                    "Year": str(y), "Month": month, "MonthNum": mn, "Shop": shop,
                    "Cashier": str(g(row, "user")).strip(),
                    "Game": clean_game_name(g(row, "game")),
                    "BetSlips": int(clean_currency_string(g(row, "bet slips"))),
                    "PaidIn": clean_currency_string(g(row, "paid in")),
                    "PaidOut": clean_currency_string(g(row, "paid out")),
                    "GrossWin": clean_currency_string(g(row, "gross win")),
                    "GWMargin": clean_currency_string(g(row, "gw margin %")),
                    "NetWin": clean_currency_string(g(row, "net win")),
                })
    if not rows:
        return pd.DataFrame(), pd.DataFrame()

    detail = pd.DataFrame(rows)
    # slip frame: per cashier (sum betslips across games for that cashier/month)
    slip_df = (detail.groupby(["Year", "Month", "MonthNum", "Shop", "Cashier"], as_index=False)
                     .agg(BetSlips=("BetSlips", "sum"), PaidIn=("PaidIn", "sum"),
                          PaidOut=("PaidOut", "sum"), NetWin=("NetWin", "sum"),
                          GWMargin=("GWMargin", "mean")))
    # game frame: per game (sum across cashiers)
    game_df = (detail.groupby(["Year", "Month", "MonthNum", "Shop", "Game"], as_index=False)
                     .agg(BetSlips=("BetSlips", "sum"), PaidIn=("PaidIn", "sum"),
                          PaidOut=("PaidOut", "sum"), GrossWin=("GrossWin", "sum")))
    game_df["PaidOutCount"] = 0
    return slip_df, game_df


# --- PRECISION HELPER FUNCTIONS ---
def clean_currency_string(val):
    if pd.isna(val) or val == '' or val == 'nan' or val == 'NaN':
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    s = re.sub(r'[R$€£]', '', s)
    s = re.sub(r'ZAR', '', s, flags=re.IGNORECASE)
    s = re.sub(r'[\s\xa0\n\r\t]', '', s)
    if s.startswith('(') and s.endswith(')'):
        s = '-' + s[1:-1]
    if ',' in s:
        comma_count = s.count(',')
        dot_count = s.count('.')
        if comma_count > 0 and dot_count > 0:
            if dot_count == 1 and comma_count == 1:
                dot_pos = s.index('.')
                comma_pos = s.index(',')
                if comma_pos > dot_pos:
                    s = s.replace('.', '')
                    s = s.replace(',', '.')
                else:
                    s = s.replace(',', '')
            elif dot_count >= 1 and comma_count >= 1:
                s = s.replace(',', '')
            elif comma_count == 1 and dot_count > 1:
                s = s.replace('.', '')
                s = s.replace(',', '.')
            else:
                s = s.replace(',', '')
        elif comma_count > 0 and dot_count == 0:
            if len(s.split(',')[1]) <= 2:
                s = s.replace(',', '.')
            else:
                s = s.replace(',', '')
    s = re.sub(r'[^\d.\-]', '', s)
    if s.count('.') > 1:
        parts = s.split('.')
        s = ''.join(parts[:-1]) + '.' + parts[-1]
    try:
        return float(s)
    except ValueError:
        match = re.search(r'[\d,\.]+', s)
        if match:
            try:
                num_str = match.group().replace(',', '')
                return float(num_str)
            except Exception:
                return 0.0
        return 0.0


def extract_date_from_filename(filename):
    month_pattern = r'(January|February|March|April|May|June|July|August|September|October|November|December)\s*(\d{4})'
    match = re.search(month_pattern, filename, re.IGNORECASE)
    if match:
        month_name = match.group(1).capitalize()
        year = int(match.group(2))
        return datetime(year, month_order.index(month_name) + 1, 1)
    month_pattern2 = r'(January|February|March|April|May|June|July|August|September|October|November|December)[_\-\s]*(\d{4})'
    match = re.search(month_pattern2, filename, re.IGNORECASE)
    if match:
        month_name = match.group(1).capitalize()
        year = int(match.group(2))
        return datetime(year, month_order.index(month_name) + 1, 1)
    month_abbr_pattern = r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s*(\d{4})'
    match = re.search(month_abbr_pattern, filename, re.IGNORECASE)
    if match:
        month_abbr = match.group(1).capitalize()
        year = int(match.group(2))
        month_map = {
            'Jan': 'January', 'Feb': 'February', 'Mar': 'March', 'Apr': 'April',
            'May': 'May', 'Jun': 'June', 'Jul': 'July', 'Aug': 'August',
            'Sep': 'September', 'Oct': 'October', 'Nov': 'November', 'Dec': 'December'
        }
        month_name = month_map.get(month_abbr, 'January')
        return datetime(year, month_order.index(month_name) + 1, 1)
    return None


def parse_jan2025_date(date_val):
    if pd.isna(date_val):
        return None
    if isinstance(date_val, (pd.Timestamp, datetime)):
        return date_val
    date_str = str(date_val).strip()
    try:
        parts = date_str.split()
        if len(parts) >= 2:
            date_parts = parts[0].split('/')
            time_parts = parts[1].split(':')
            if len(date_parts) == 3 and len(time_parts) >= 2:
                day, month, year = int(date_parts[0]), int(date_parts[1]), int(date_parts[2])
                if year < 100:
                    year += 2000
                hour, minute = int(time_parts[0]), int(time_parts[1])
                second = int(time_parts[2]) if len(time_parts) > 2 else 0
                return pd.Timestamp(year=year, month=month, day=day, hour=hour, minute=minute, second=second)
    except Exception:
        pass
    try:
        return pd.to_datetime(date_str, format='%d/%m/%y %H:%M:%S', errors='coerce')
    except Exception:
        try:
            return pd.to_datetime(date_str, dayfirst=True, errors='coerce')
        except Exception:
            return None


def find_deposit_column(df):
    deposit_keywords = ['paid in sum', 'paidin', 'paid in', 'deposits', 'deposit',
                        'cash in', 'cashin', 'payment', 'paid_sum', 'paid sum', 'paid-in', 'paid_in']
    for col in df.columns:
        if str(col).lower().strip() in deposit_keywords:
            return col
    for col in df.columns:
        col_lower = str(col).lower().strip()
        for keyword in deposit_keywords:
            if keyword in col_lower:
                return col
    for col in df.columns:
        col_lower = str(col).lower().strip()
        if 'paid' in col_lower or 'deposit' in col_lower:
            return col
    return None


def find_ggr_column(df):
    ggr_keywords = ['gross win', 'grosswin', 'ggr', 'gross revenue', 'grossrevenue', 'gross', 'win', 'revenue']
    for col in df.columns:
        if str(col).lower().strip() in ggr_keywords:
            return col
    for col in df.columns:
        col_lower = str(col).lower().strip()
        for keyword in ggr_keywords:
            if keyword in col_lower:
                return col
    for col in df.columns:
        col_lower = str(col).lower().strip()
        if 'gross' in col_lower or 'win' in col_lower:
            return col
    return None


def clean_game_name(val):
    if pd.isna(val):
        return "Unknown Game"
    if isinstance(val, (int, float)):
        return str(int(val)) if val == int(val) else str(val)
    s = str(val).strip()
    s = re.sub(r'\s+', ' ', s)
    s = re.sub(r'(?i)\bbetgames\b', 'BetGames', s)
    s = re.sub(r'(?i)\bskypilot\b', 'SkyPilot', s)
    return s


def enforce_schema(df):
    if df is None or df.empty:
        return None
    df = df.loc[:, ~df.columns.duplicated()]
    if 'Shop' in df.columns:
        df['Shop'] = df['Shop'].astype(str).str.strip().replace({'nan': None, 'NaN': None, 'None': None, '': None})
        df['Shop'] = df['Shop'].ffill()
        df['Shop'] = df['Shop'].replace({'Potch': 'Potchefstroom'})
    else:
        df['Shop'] = 'Unknown'
    if 'Game' in df.columns:
        df['Game'] = df['Game'].apply(clean_game_name)
    else:
        df['Game'] = 'Unknown Game'
    for col in TARGET_COLS:
        if col not in df.columns:
            df[col] = 0.0 if col not in ['Shop', 'Game', 'Year', 'Month'] else "Unknown"
    num_cols = ['Deposits', 'Paid Out Sum', 'GGR', 'GW Margin %', 'Net Win', 'Net Win Margin']
    for col in num_cols:
        if col in df.columns:
            df[col] = df[col].apply(clean_currency_string)
        else:
            df[col] = 0.0
    return df[TARGET_COLS].copy()


def is_per_user_sheet(df_raw):
    if df_raw.empty:
        return False
    first_row = [str(x).strip().lower() for x in df_raw.iloc[0].values]
    return 'game' in first_row and 'shop' in first_row and 'user' in first_row


def is_per_user_csv(df):
    """Detect a per-user slip-summary CSV: has Shop + User + Paid In, but no Game."""
    cols = [str(c).lower().strip() for c in df.columns]
    has_shop = 'shop' in cols
    has_user = 'user' in cols
    has_paid_in = any(c in ('paid in', 'paidin', 'paid in sum') for c in cols)
    has_game = 'game' in cols
    return has_shop and has_user and has_paid_in and not has_game


def process_per_user_csv(df, file_date):
    """Aggregate a per-user slip-summary CSV to one row per branch.

    Deposits  <- sum of 'Paid In'
    GGR       <- sum of 'Net Win'   (Paid In - Paid Out - adjustments)
    Paid Out  <- sum of 'Paid Out'
    """
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    def col(*names):
        for n in names:
            for c in df.columns:
                if str(c).lower().strip() == n:
                    return c
        return None

    shop_col = col('shop')
    paid_in_col = col('paid in', 'paidin', 'paid in sum')
    net_win_col = col('net win')
    paid_out_col = col('paid out')

    df['_dep'] = df[paid_in_col].apply(clean_currency_string) if paid_in_col else 0.0
    df['_ggr'] = df[net_win_col].apply(clean_currency_string) if net_win_col else 0.0
    df['_po'] = df[paid_out_col].apply(clean_currency_string) if paid_out_col else 0.0
    df['Shop'] = df[shop_col].astype(str).str.strip() if shop_col else 'Unknown'
    df['Shop'] = df['Shop'].replace({'Potch': 'Potchefstroom'})
    df = df[df['Shop'].isin(BRANCHES)]
    if df.empty:
        return None

    grouped = df.groupby('Shop', as_index=False).agg(
        Deposits=('_dep', 'sum'),
        GGR=('_ggr', 'sum'),
        **{'Paid Out Sum': ('_po', 'sum')}
    )
    grouped['Game'] = 'All Games'
    grouped['GW Margin %'] = 0.0
    grouped['Net Win'] = grouped['GGR']
    grouped['Net Win Margin'] = 0.0
    grouped['Year'] = str(file_date.year)
    grouped['Month'] = file_date.strftime('%B')
    grouped['MonthNum'] = file_date.month
    return grouped


def find_summary_header_rows(df_raw):
    header_rows = []
    for i, row in df_raw.iterrows():
        row_clean = [str(x).strip().lower() for x in row.values]
        if 'shop' in row_clean and 'game' in row_clean:
            header_rows.append(i)
    return header_rows


def parse_summary_block(df_raw, header_idx, block_end_idx, source_name):
    header_row = [str(c).strip() for c in df_raw.iloc[header_idx].values]
    block = df_raw.iloc[header_idx + 1: block_end_idx].copy()
    block.columns = header_row
    block = block.loc[:, ~block.columns.duplicated()]
    shop_col = next((c for c in block.columns if str(c).lower().strip() == 'shop'), None)
    if not shop_col:
        return None
    block['Shop'] = block[shop_col]
    gross_col = find_ggr_column(block)
    block['GGR'] = block[gross_col] if gross_col else 0.0
    deposit_col = find_deposit_column(block)
    block['Deposits'] = block[deposit_col] if deposit_col else 0.0
    date_col = next((c for c in block.columns if 'firstslip' in str(c).lower().replace(' ', '') or 'date' in str(c).lower()), None)
    if date_col:
        if 'jan' in source_name.lower():
            block['First Slip Issued'] = block[date_col].apply(parse_jan2025_date)
        else:
            block['First Slip Issued'] = pd.to_datetime(block[date_col], errors='coerce', dayfirst=True)
        block['First Slip Issued'] = block['First Slip Issued'].ffill()
        block = block.dropna(subset=['First Slip Issued'])
        block['Year'] = block['First Slip Issued'].dt.year.astype(int).astype(str)
        block['Month'] = block['First Slip Issued'].dt.strftime('%B')
        block['MonthNum'] = block['First Slip Issued'].dt.month
    else:
        return None
    block = block[block['Shop'].astype(str).str.strip().isin(BRANCHES)]
    return block


def process_excel_dataframe(df_raw, source_name):
    header_rows = find_summary_header_rows(df_raw)
    if not header_rows:
        return None
    parsed_blocks = []
    for i, header_idx in enumerate(header_rows):
        block_end_idx = header_rows[i + 1] if i + 1 < len(header_rows) else len(df_raw)
        block = parse_summary_block(df_raw, header_idx, block_end_idx, source_name)
        if block is not None and not block.empty:
            parsed_blocks.append(block)
    if not parsed_blocks:
        return None
    return pd.concat(parsed_blocks, ignore_index=True)


# --- LOAD DATA FUNCTIONS ---
@st.cache_data
def load_data(uploaded_files):
    all_data = []
    for file in uploaded_files:
        try:
            filename = file.name.lower()
            file_date = extract_date_from_filename(filename)
            if filename.endswith('.csv'):
                df = pd.read_csv(file)
                df.columns = [str(c).strip() for c in df.columns]

                # Per-user slip-summary CSV: aggregate to one row per branch
                if is_per_user_csv(df) and file_date is not None:
                    processed = process_per_user_csv(df, file_date)
                    df_clean = enforce_schema(processed)
                    if df_clean is not None:
                        all_data.append(df_clean)
                    continue

                deposit_col = find_deposit_column(df)
                if deposit_col:
                    df = df.rename(columns={deposit_col: 'Deposits'})
                ggr_col = find_ggr_column(df)
                if ggr_col:
                    df = df.rename(columns={ggr_col: 'GGR'})
                if 'Deposits' not in df.columns:
                    if 'Paid In Sum' in df.columns:
                        df = df.rename(columns={'Paid In Sum': 'Deposits'})
                    else:
                        df['Deposits'] = 0.0
                if 'GGR' not in df.columns:
                    if 'Gross Win' in df.columns:
                        df = df.rename(columns={'Gross Win': 'GGR'})
                    else:
                        df['GGR'] = 0.0
                df['Date'] = file_date
                if not df.empty and file_date:
                    df['Year'] = str(file_date.year)
                    df['Month'] = file_date.strftime('%B')
                    df['MonthNum'] = file_date.month
                df_clean = enforce_schema(df)
                if df_clean is not None:
                    all_data.append(df_clean)
            elif filename.endswith(('.xls', '.xlsx')):
                xl = pd.ExcelFile(file)
                for sheet in xl.sheet_names:
                    df_raw = pd.read_excel(file, sheet_name=sheet, header=None)
                    if is_per_user_sheet(df_raw):
                        continue
                    processed_df = process_excel_dataframe(df_raw, f"{file.name} - {sheet}")
                    df_clean = enforce_schema(processed_df)
                    if df_clean is not None:
                        all_data.append(df_clean)
        except Exception as e:
            st.error(f"Error loading {file.name}: {e}")
    if all_data:
        return pd.concat(all_data, ignore_index=True)
    return pd.DataFrame()


@st.cache_data
def load_historical_from_folder():
    all_data = []
    file_count = 0
    excel_files = glob.glob(os.path.join(HISTORICAL_FOLDER, "*.xlsx")) + \
                  glob.glob(os.path.join(HISTORICAL_FOLDER, "*.xls"))
    if not excel_files:
        return pd.DataFrame(), 0
    for file_path in excel_files:
        try:
            filename = os.path.basename(file_path)
            xl = pd.ExcelFile(file_path)
            for sheet in xl.sheet_names:
                df_raw = pd.read_excel(file_path, sheet_name=sheet, header=None)
                if is_per_user_sheet(df_raw):
                    continue
                processed_df = process_excel_dataframe(df_raw, f"{filename} - {sheet}")
                df_clean = enforce_schema(processed_df)
                if df_clean is not None:
                    all_data.append(df_clean)
                    file_count += 1
        except Exception as e:
            st.warning(f"⚠️ Could not load {file_path}: {str(e)}")
    if all_data:
        return pd.concat(all_data, ignore_index=True), file_count
    return pd.DataFrame(), 0


@st.cache_data
def load_uploaded_csvs_from_folder():
    all_data = []
    file_count = 0
    csv_files = glob.glob(os.path.join(UPLOAD_FOLDER, "*.csv"))
    if not csv_files:
        return pd.DataFrame(), 0
    for file_path in csv_files:
        try:
            filename = os.path.basename(file_path)
            file_date = extract_date_from_filename(filename)
            if file_date is None:
                continue
            # Peek at header only — skip giant raw per-slip dumps before loading them.
            head = pd.read_csv(file_path, nrows=0)
            head.columns = [str(c).strip() for c in head.columns]
            if is_raw_betslip_csv(head):
                continue
            df = pd.read_csv(file_path)
            df.columns = [str(c).strip() for c in df.columns]

            # Cash Operations exports are per game×cashier. Aggregate to one row per
            # branch (sum Gross Win -> GGR, sum Paid In Sum -> Deposits) so the month
            # total is correct. The merge stage below drops any month that already has
            # a slip figure, so a month is never counted twice.
            if is_cash_ops_csv(df):
                sub = df.copy()
                shop_c = _col(sub, 'shop')
                sub['Shop'] = sub[shop_c].astype(str).str.strip().replace({'Potch': 'Potchefstroom'})
                sub = sub[sub['Shop'].isin(BRANCHES)]
                if sub.empty:
                    continue
                gin = _col(sub, 'paid in sum')
                gw = _col(sub, 'gross win')
                po = _col(sub, 'paid out sum')
                sub['_dep'] = sub[gin].apply(clean_currency_string) if gin else 0.0
                sub['_ggr'] = sub[gw].apply(clean_currency_string) if gw else 0.0
                sub['_po'] = sub[po].apply(clean_currency_string) if po else 0.0
                game_col = _col(sub, 'game')
                grouped = sub.groupby(['Shop', game_col], as_index=False).agg(
                    Deposits=('_dep', 'sum'), GGR=('_ggr', 'sum'),
                    **{'Paid Out Sum': ('_po', 'sum')})
                grouped = grouped.rename(columns={game_col: 'Game'})
                grouped['Game'] = grouped['Game'].apply(clean_game_name)
                grouped['GW Margin %'] = 0.0
                grouped['Net Win'] = grouped['GGR']
                grouped['Net Win Margin'] = 0.0
                grouped['Year'] = str(file_date.year)
                grouped['Month'] = file_date.strftime('%B')
                grouped['MonthNum'] = file_date.month
                df_clean = enforce_schema(grouped)
                if df_clean is not None:
                    all_data.append(df_clean)
                    file_count += 1
                continue

            # Per-user slip-summary CSV: aggregate to one row per branch
            if is_per_user_csv(df):
                processed = process_per_user_csv(df, file_date)
                df_clean = enforce_schema(processed)
                if df_clean is not None:
                    all_data.append(df_clean)
                    file_count += 1
                continue

            deposit_col = find_deposit_column(df)
            if deposit_col:
                df = df.rename(columns={deposit_col: 'Deposits'})
            elif 'Paid In Sum' in df.columns:
                df = df.rename(columns={'Paid In Sum': 'Deposits'})
            elif 'Deposits' not in df.columns:
                df['Deposits'] = 0.0
            ggr_col = find_ggr_column(df)
            if ggr_col:
                df = df.rename(columns={ggr_col: 'GGR'})
            elif 'Gross Win' in df.columns:
                df = df.rename(columns={'Gross Win': 'GGR'})
            elif 'GGR' not in df.columns:
                df['GGR'] = 0.0
            shop_col = None
            for col in df.columns:
                if str(col).lower().strip() in ['shop', 'branch', 'store', 'location']:
                    shop_col = col
                    break
            if shop_col:
                df = df.rename(columns={shop_col: 'Shop'})
            elif 'Shop' not in df.columns:
                df['Shop'] = 'Malvern'
            df['Date'] = file_date
            if not df.empty:
                df['Year'] = str(file_date.year)
                df['Month'] = file_date.strftime('%B')
                df['MonthNum'] = file_date.month
            df_clean = enforce_schema(df)
            if df_clean is not None:
                all_data.append(df_clean)
                file_count += 1
        except Exception as e:
            st.warning(f"⚠️ Could not load {file_path}: {str(e)}")
    if all_data:
        return pd.concat(all_data, ignore_index=True), file_count
    return pd.DataFrame(), 0


def ensure_numeric(df, columns):
    for col in columns:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
            df[col] = df[col].fillna(0)
    return df


# --- APP LAYOUT ---
st.title("Playbet Dashboard")
_boot.empty()

st.sidebar.header("📤 Upload Data")
st.sidebar.caption("Two separate uploaders — put each file in its own box.")

# ---- Uploader 1: SLIP SUMMARY (per-cashier) ----
st.sidebar.markdown("**1 · Slip Summary** (per cashier)")
slip_files = st.sidebar.file_uploader(
    "Slip Summary CSV(s)",
    accept_multiple_files=True, type=["csv"], key="slip_uploader",
    help="Per-cashier slip export. Feeds the cashier comparison AND the GGR/Deposits dashboard."
)

# ---- Uploader 2: CASH OPERATIONS (per game) ----
st.sidebar.markdown("**2 · Cash Operations** (per game)")
cash_files = st.sidebar.file_uploader(
    "Cash Operations CSV(s)",
    accept_multiple_files=True, type=["csv"], key="cash_uploader",
    help="Per-game cash-ops export. Feeds the game comparison only — kept out of the GGR chart."
)

st.sidebar.info("💡 Filenames must include month and year (e.g., 'May 2026.csv').")

# Save uploads. Each uploader routes to ONLY the correct destination, so a
# Cash Operations file can never be summed into the GGR/Deposits chart (which
# was the cause of the inflated May bar).
def _process_uploads(files, kind):
    """kind: 'slip' or 'cash'. Returns (saved_names, dup_names)."""
    from io import BytesIO
    saved, dups = [], []
    for file in files:
        try:
            file_date = extract_date_from_filename(file.name.lower())
            if file_date is None:
                st.sidebar.warning(f"⚠️ {file.name}: no month/year in filename — skipped.")
                continue
            file_bytes = file.getvalue()
            raw = pd.read_csv(BytesIO(file_bytes))
            raw.columns = [str(c).strip() for c in raw.columns]

            outcomes = []
            if kind == "cash":
                # Cash Operations -> GAME detail table ONLY. Never the GGR summary.
                if is_cash_ops_csv(raw):
                    outcomes.append(save_game_rows_to_neon(raw, file_date, file.name))
                else:
                    st.sidebar.warning(f"⚠️ {file.name}: doesn't look like a Cash Operations export.")
                    continue
            else:  # slip
                if is_per_user_csv(raw):
                    # cashier detail for the comparison
                    outcomes.append(save_slip_rows_to_neon(raw, file_date, file.name))
                    # AND the existing GGR/Deposits summary dashboard
                    try:
                        file.seek(0)
                    except Exception:
                        pass
                    parsed = load_data([file])
                    if parsed is not None and not parsed.empty:
                        outcomes.append(save_uploaded_rows_to_neon(parsed, file.name))
                else:
                    st.sidebar.warning(f"⚠️ {file.name}: doesn't look like a Slip Summary export.")
                    continue

            if not outcomes:
                st.sidebar.warning(f"⚠️ {file.name}: parsed to 0 rows.")
            elif "saved" in outcomes:
                saved.append(file.name)
            elif all(o == "duplicate" for o in outcomes):
                dups.append(file.name)
        except Exception as e:
            st.sidebar.warning(f"Could not process {file.name}: {e}")
    return saved, dups


if slip_files or cash_files:
    if _engine is None:
        st.sidebar.error("❌ No database connection — upload won't persist.")
    all_saved, all_dups = [], []
    if slip_files:
        s, d = _process_uploads(slip_files, "slip")
        all_saved += s
        all_dups += d
    if cash_files:
        s, d = _process_uploads(cash_files, "cash")
        all_saved += s
        all_dups += d
    if all_dups:
        st.sidebar.error(f"🚫 Already uploaded — blocked: {', '.join(all_dups)}")
    if all_saved:
        st.sidebar.success(f"✅ Saved {len(all_saved)} file(s) to database")
        st.cache_data.clear()
        st.rerun()

# ---- Maintenance: clear ingested data so you can re-upload cleanly ----
with st.sidebar.expander("🧹 Maintenance"):
    st.caption("Reload re-reads the files in the repo folders (use after adding files "
               "to the uploads/ folder). Clear deletes uploaded rows from the database.")
    if st.button("Reload data from files"):
        st.cache_data.clear()
        st.success("Cache cleared — re-reading all files.")
        st.rerun()

    if st.button("Backfill missing months into comparison"):
        # For any 2026 month that has a branch-level GGR summary in
        # dashboard_uploaded_rows but no per-game detail in dashboard_game_rows
        # (e.g. June, July), copy the branch totals in as one 'All Games' row per
        # branch so those months appear in the games comparison. Branch-level only —
        # per-game detail was never stored for these months.
        if _engine is None:
            st.error("No database connection.")
        else:
            try:
                with _engine.begin() as c:
                    summary_months = c.execute(text(
                        "SELECT DISTINCT year, month, monthnum FROM dashboard_uploaded_rows WHERE year='2026'"
                    )).fetchall()
                    detail = set((r[0], r[1]) for r in c.execute(text(
                        "SELECT DISTINCT year, month FROM dashboard_game_rows WHERE year='2026'"
                    )).fetchall())
                    todo = [(y, m, mn) for (y, m, mn) in summary_months if (y, m) not in detail]
                    if not todo:
                        st.info("Nothing to backfill — all 2026 summary months already have detail.")
                    else:
                        done = []
                        for (year, month, monthnum) in todo:
                            rows = c.execute(text("""
                                SELECT shop, SUM(COALESCE(ggr,0)) AS ggr,
                                       SUM(COALESCE(deposits,0)) AS paid_in,
                                       SUM(COALESCE(paid_out_sum,0)) AS paid_out
                                FROM dashboard_uploaded_rows
                                WHERE year=:y AND month=:m GROUP BY shop
                            """), {"y": year, "m": month}).fetchall()
                            c.execute(text("""
                                DELETE FROM dashboard_game_rows
                                WHERE year=:y AND month=:m AND game='All Games'
                            """), {"y": year, "m": month})
                            for shop, ggr, paid_in, paid_out in rows:
                                c.execute(text("""
                                    INSERT INTO dashboard_game_rows
                                        (source_file, shop, game, paid_in_count, paid_in_sum,
                                         paid_out_count, paid_out_sum, gross_win, year, month, monthnum)
                                    VALUES ('backfill_summary', :shop, 'All Games', 0, :pin,
                                            0, :pout, :gw, :y, :m, :mn)
                                """), {"shop": shop, "pin": float(paid_in or 0),
                                       "pout": float(paid_out or 0), "gw": float(ggr or 0),
                                       "y": year, "m": month, "mn": int(monthnum)})
                            done.append(f"{month} {year}")
                        st.cache_data.clear()
                        st.success("Backfilled: " + ", ".join(done) + ". These show as 'All Games' per branch.")
                        st.rerun()
            except Exception as e:
                st.error(f"Backfill failed: {e}")
    if st.button("Clear uploaded & comparison data"):
        if _engine is not None:
            try:
                with _engine.begin() as c:
                    c.execute(text("DELETE FROM dashboard_uploaded_rows"))
                    c.execute(text("DELETE FROM dashboard_slip_rows"))
                    c.execute(text("DELETE FROM dashboard_game_rows"))
                    c.execute(text("DELETE FROM dashboard_upload_hashes"))
                st.cache_data.clear()
                st.success("Cleared. Re-upload your Slip Summary and Cash Operations files.")
                st.rerun()
            except Exception as e:
                st.error(f"Could not clear: {e}")
        else:
            st.error("No database connection.")

# --- SIDEBAR: FILTERS ---
st.sidebar.divider()
st.sidebar.header("⏳ Filters")

# --- MAIN RUN LOGIC ---
historical_df, historical_file_count = load_historical_from_folder()

# Uploaded data comes from two real sources:
#   1. CSVs committed in the local uploads/ folder (historical months)
#   2. Rows saved to Neon (newer uploads, survive restarts)
# Both are read, then deduped by Month+Year+Shop so a month present in BOTH
# is counted once (Neon wins), preventing the doubled-figures problem.
folder_uploaded_df, _ = load_uploaded_csvs_from_folder()
neon_uploaded_df = load_uploaded_rows_from_neon()
manual_df = load_manual_entries_from_neon()

if folder_uploaded_df.empty:
    uploaded_df = neon_uploaded_df
elif neon_uploaded_df.empty:
    uploaded_df = folder_uploaded_df
else:
    # Drop any folder rows whose Month+Year+Shop already exists in Neon
    neon_keys = set(
        zip(neon_uploaded_df['Year'].astype(str),
            neon_uploaded_df['Month'].astype(str),
            neon_uploaded_df['Shop'].astype(str))
    )
    mask = folder_uploaded_df.apply(
        lambda r: (str(r['Year']), str(r['Month']), str(r['Shop'])) not in neon_keys,
        axis=1
    )
    folder_unique = folder_uploaded_df[mask]
    uploaded_df = pd.concat([neon_uploaded_df, folder_unique], ignore_index=True)

df_parts = []
if not historical_df.empty:
    df_parts.append(historical_df)
if not uploaded_df.empty:
    df_parts.append(uploaded_df)
if manual_df is not None and not manual_df.empty:
    df_parts.append(manual_df)

if df_parts:
    df = pd.concat(df_parts, ignore_index=True)
    # Safety net: never show 2026 May/June/July (other years keep those months)
    df = df[~df.apply(lambda r: is_blocked(r['Month'], r['Year']), axis=1)]
    numeric_cols = ['GGR', 'Deposits', 'Paid Out Sum', 'GW Margin %', 'Net Win', 'Net Win Margin']
    df = ensure_numeric(df, numeric_cols)
    if 'Game' in df.columns:
        df['Game'] = df['Game'].fillna('Unknown Game').astype(str)

    if not df.empty:
        available_months = sorted(df['Month'].unique(), key=lambda m: month_order.index(m) if m in month_order else 0)
        available_years = sorted(df['Year'].unique())

        selected_year = st.sidebar.selectbox("Select Year:", ["All Time"] + available_years)
        selected_month = st.sidebar.selectbox("Select Month:", ["All Months"] + available_months) if selected_year != "All Time" else "All Months"
        nav_options = ["All Branches Dashboard"] + BRANCHES
        selected_view = st.sidebar.radio("Select Branch Analysis:", nav_options)

        df_filtered = df if selected_view == "All Branches Dashboard" else df[df['Shop'] == selected_view]
        if selected_year != "All Time":
            df_filtered = df_filtered[df_filtered['Year'] == selected_year]
        if selected_month != "All Months":
            df_filtered = df_filtered[df_filtered['Month'] == selected_month]

        df_filtered = ensure_numeric(df_filtered, ['GGR', 'Deposits'])
        if 'Game' in df_filtered.columns:
            df_filtered['Game'] = df_filtered['Game'].fillna('Unknown Game').astype(str)

        if not df_filtered.empty and len(df_filtered) > 0:
            total_ggr = df_filtered['GGR'].sum()
            total_deposits = df_filtered['Deposits'].sum()
            try:
                game_ggr = df_filtered.groupby('Game')['GGR'].sum()
                top_game = game_ggr.idxmax() if not game_ggr.empty else "N/A"
            except Exception:
                top_game = "N/A"
        else:
            total_ggr = 0.0
            total_deposits = 0.0
            top_game = "N/A"

        yoy = 0.0
        if selected_year == "All Time" and not df_filtered.empty:
            years = sorted(df_filtered['Year'].unique())
            if len(years) >= 2:
                curr = df_filtered[df_filtered['Year'] == years[-1]]['GGR'].sum()
                prev = df_filtered[df_filtered['Year'] == years[-2]]['GGR'].sum()
                yoy = ((curr - prev) / prev) * 100 if prev != 0 else 0

        st.subheader(f"{selected_view} Performance")
        c1, c2, c3, c4 = st.columns(4)

        def render_metric_card(col, label, value_str, color=None):
            value_style = "font-size:1.75rem; font-weight:600; line-height:1.2;"
            if color:
                value_style += f" color:{color};"
            col.markdown(
                f"""
                <div style="display:flex; flex-direction:column; gap:0.25rem;">
                    <span style="font-size:0.875rem; color:rgba(49,51,63,0.6);">{label}</span>
                    <span style="{value_style}">{value_str}</span>
                </div>
                """,
                unsafe_allow_html=True
            )

        render_metric_card(c1, "Total GGR", f"R {total_ggr:,.2f}")
        render_metric_card(c2, "Total Deposits (Paid In)", f"R {total_deposits:,.2f}")
        # YoY Growth: green when positive, red when negative.
        if selected_year == "All Time":
            yoy_color = "#16A34A" if yoy > 0 else ("#C0392B" if yoy < 0 else None)
            render_metric_card(c3, "YoY Growth", f"{yoy:+.1f}%", color=yoy_color)
        else:
            render_metric_card(c3, "YoY Growth", "N/A (Filtered)")
        render_metric_card(c4, "Top Performer", top_game)
        st.divider()

        st.subheader("GGR: Multi-Year Stacked Comparison")
        chart_data = df_filtered.groupby(['MonthNum', 'Month', 'Year'])['GGR'].sum().reset_index().sort_values('MonthNum')
        if not chart_data.empty:
            fig = px.bar(chart_data, x='Month', y='GGR', color='Year', barmode='stack', category_orders={"Month": month_order}, color_discrete_map=YEAR_COLORS)
            fig.update_layout(xaxis_title=None, yaxis_title="Gross Gaming Revenue (ZAR)")
            fig.update_traces(hovertemplate="<b>%{x} %{fullData.name}</b><br>GGR: R %{y:,.2f}<extra></extra>")
            fig.update_layout(yaxis_tickformat=",.0f")
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.warning("No data available for GGR chart")

        st.subheader("Deposits (Paid In): Multi-Year Stacked Comparison")
        chart_data_deposits = df_filtered.groupby(['MonthNum', 'Month', 'Year'])['Deposits'].sum().reset_index().sort_values('MonthNum')
        if not chart_data_deposits.empty:
            fig_deposits = px.bar(chart_data_deposits, x='Month', y='Deposits', color='Year', barmode='stack', category_orders={"Month": month_order}, color_discrete_map=DEPOSIT_YEAR_COLORS)
            fig_deposits.update_layout(xaxis_title=None, yaxis_title="Deposits / Paid In (ZAR)")
            fig_deposits.update_traces(hovertemplate="<b>%{x} %{fullData.name}</b><br>Deposits: R %{y:,.2f}<extra></extra>")
            fig_deposits.update_layout(yaxis_tickformat=",.0f")
            st.plotly_chart(fig_deposits, use_container_width=True)
        else:
            st.warning("No data available for Deposits chart")

        # (Game Revenue Analysis removed — its GGR-by-branch-and-year and
        # month-to-month GGR tables are replaced by the Quarterly Comparison below,
        # which shows games per branch and betslips by month for 2026.)
        st.divider()

        # =============================================================
        # QUARTERLY COMPARISON — games & cashiers (betslips, GWM%, best/worst month)
        # Runs on the full-detail slip/game rows kept in Neon, scoped to a
        # year + quarter the user picks, honouring the branch selected above.
        # =============================================================
        st.header("📈 Quarterly Comparison")

        slip_full = load_slip_rows_from_neon()
        game_full = load_game_rows_from_neon()

        # Also read detail straight from the repo uploads/ folder, so months already
        # committed there appear in the comparison without re-uploading. Neon wins on
        # any (Year, Month, Shop) overlap to avoid double-counting.
        folder_slip, folder_game = load_comparison_detail_from_folder()
        hist_slip, hist_game = load_comparison_detail_from_historical()

        def _merge_detail(neon_df, folder_df, keys):
            if folder_df is None or folder_df.empty:
                return neon_df
            if neon_df is None or neon_df.empty:
                return folder_df
            existing = set(zip(*[neon_df[k].astype(str) for k in keys]))
            mask = folder_df.apply(lambda r: tuple(str(r[k]) for k in keys) not in existing, axis=1)
            return pd.concat([neon_df, folder_df[mask]], ignore_index=True)

        slip_full = _merge_detail(slip_full, folder_slip, ["Year", "Month", "Shop"])
        game_full = _merge_detail(game_full, folder_game, ["Year", "Month", "Shop"])
        slip_full = _merge_detail(slip_full, hist_slip, ["Year", "Month", "Shop"])
        game_full = _merge_detail(game_full, hist_game, ["Year", "Month", "Shop"])

        if slip_full.empty and game_full.empty:
            st.info("💡 Upload a Slip Summary and/or Cash Operations CSV to build the quarterly comparison. "
                    "These carry per-cashier and per-game betslip detail the comparison needs.")
        else:
            # This section is scoped to 2026 only, driven by the Slip Summary +
            # Cash Operations uploads. Only the quarter is selectable.
            comp_year = "2026"
            # keep only 2026 detail
            slip_full = slip_full[slip_full["Year"].astype(str) == "2026"] if not slip_full.empty else slip_full
            game_full = game_full[game_full["Year"].astype(str) == "2026"] if not game_full.empty else game_full

            if (slip_full is None or slip_full.empty) and (game_full is None or game_full.empty):
                st.info("💡 No 2026 comparison data yet. Upload the Slip Summary and Cash Operations files.")
                _has_2026 = False
            else:
                _has_2026 = True

            if _has_2026:
                st.caption("Showing 2026 only.")
                comp_quarter = st.selectbox("Comparison Quarter:", list(cmp.QUARTERS.keys()), index=0)

                comp_branch = None if selected_view == "All Branches Dashboard" else selected_view

                slip_s = cmp.scope(slip_full, comp_year, comp_quarter, comp_branch)
                game_s = cmp.scope(game_full, comp_year, comp_quarter, comp_branch)

                discontinued = set()

                scope_label = f"{comp_year} {comp_quarter}" + (f" · {comp_branch}" if comp_branch else " · All Branches")
                st.caption(f"Showing: **{scope_label}**")

                if (slip_s is None or slip_s.empty) and (game_s is None or game_s.empty):
                    st.warning("No comparison data for this quarter/branch selection.")
                else:
                    comp_tables = {
                        "Best-Worst Month":        cmp.best_worst_month(slip_s, game_s, comp_quarter),
                        "Games Betslips by Month": cmp.games_betslip_h2h(game_s, comp_quarter),
                        "Games Increase-Decrease": cmp.games_increase_decrease(game_s, comp_quarter),
                        "Games per Branch":        cmp.games_per_branch(game_s),
                        "Games GWM":               cmp.games_gwm(game_s, comp_quarter),
                        "Cashier Betslips by Month": cmp.cashier_betslip_h2h(slip_s, comp_quarter),
                        "Cashier GWM":             cmp.cashier_gwm(slip_s, comp_quarter),
                    }

                    def show(df, caption=None):
                        if df is None or df.empty:
                            st.info("No data for this selection.")
                            return
                        if caption:
                            st.caption(caption)
                        # In each row, colour the highest figure green and the lowest
                        # red, across the monthly columns only. Covers bare month
                        # columns (betslips) and monthly "... % January" GWM columns,
                        # but never the Quarter/Total columns.
                        month_cols = [c for c in df.columns
                                      if (c in month_order or
                                          (any(m in str(c) for m in month_order) and "Quarter" not in str(c)))
                                      and "Total" not in str(c)]
                        if month_cols and len(month_cols) >= 2:
                            def _row_hilite(row):
                                styles = pd.Series("", index=row.index)
                                vals = pd.to_numeric(row[month_cols], errors="coerce")
                                if vals.notna().any() and vals.max() != vals.min():
                                    hi = vals.idxmax()
                                    lo = vals.idxmin()
                                    styles[hi] = "background-color: #16A34A; color: white; font-weight: bold;"
                                    styles[lo] = "background-color: #C0392B; color: white; font-weight: bold;"
                                return styles
                            styled = df.style.apply(_row_hilite, axis=1)
                            st.dataframe(styled, use_container_width=True, hide_index=True)
                        else:
                            st.dataframe(df, use_container_width=True, hide_index=True)

                    tab_sum, tab_games, tab_cash = st.tabs(
                        ["Summary", "Games", "Cashiers"])

                    with tab_sum:
                        st.subheader("Busiest and least-busy month, best and worst value month")
                        st.caption("Busiest / least-busy is ranked by total betslip count. "
                                   "Best / worst value is ranked by total Gross Win for the month.")
                        bw = comp_tables["Best-Worst Month"]
                        show(bw)
                        if not bw.empty and "Betslips" in bw.columns:
                            fig_bw = px.bar(bw, x="Month", y="Betslips", color="Month",
                                            template="plotly_white")
                            fig_bw.update_layout(showlegend=False, yaxis_tickformat=",.0f",
                                                 margin=dict(t=10, b=10))
                            st.plotly_chart(fig_bw, use_container_width=True)

                    with tab_games:
                        st.subheader("Betslip count by game, month by month")
                        st.caption("Each game's betslip count for each month of the quarter, side by side.")
                        show(comp_tables["Games Betslips by Month"])

                        st.subheader("Betslip increases and decreases by game")
                        st.caption("Change in betslip count from the first month to the last month of the quarter.")
                        show(comp_tables["Games Increase-Decrease"])

                        st.subheader("Games per branch")
                        st.caption("Betslips, Paid In, Gross Win and Gross Win Margin for each game in each branch.")
                        show(comp_tables["Games per Branch"])

                        st.subheader("Gross Win Margin percentage by game")
                        show(comp_tables["Games GWM"],
                             "Gross Win Margin % = Gross Win ÷ Paid In. Shown per month and for the full quarter.")

                    with tab_cash:
                        st.subheader("Betslip count by cashier, month by month")
                        st.caption("Each cashier's betslip count for each month of the quarter, side by side.")
                        show(comp_tables["Cashier Betslips by Month"])

                        st.subheader("Gross Win Margin percentage by cashier")
                        show(comp_tables["Cashier GWM"],
                             "Gross Win Margin % per month. The quarter column is weighted by Paid In, "
                             "not a simple average of the monthly percentages.")

                    st.divider()

                    # Branch-by-branch workbook: build each branch's tables from the full
                    # (unscoped-by-branch) frames, so every branch gets its own sheet.
                    from comparison_excel import build_branch_by_branch_workbook

                    # Frames scoped to the year+quarter but NOT to a single branch.
                    slip_q = cmp.scope(slip_full, comp_year, comp_quarter, None)
                    game_q = cmp.scope(game_full, comp_year, comp_quarter, None)
                    if discontinued and game_q is not None and not game_q.empty:
                        game_q = game_q[~game_q["Game"].isin(discontinued)].copy()

                    branches_present = sorted(
                        set(([] if game_q is None or game_q.empty else list(game_q["Shop"].unique())) +
                            ([] if slip_q is None or slip_q.empty else list(slip_q["Shop"].unique())))
                    )
                    per_branch_tables = {}
                    for b in branches_present:
                        gb = game_q[game_q["Shop"] == b] if game_q is not None and not game_q.empty else game_q
                        sb = slip_q[slip_q["Shop"] == b] if slip_q is not None and not slip_q.empty else slip_q
                        per_branch_tables[b] = {
                            "Games Betslips by Month": cmp.games_betslip_h2h(gb, comp_quarter),
                            "Games GWM": cmp.games_gwm(gb, comp_quarter),
                            "Cashier Betslips by Month": cmp.cashier_betslip_h2h(sb, comp_quarter),
                            "Cashier GWM": cmp.cashier_gwm(sb, comp_quarter),
                        }

                    safe = f"{comp_year} {comp_quarter}".replace(" ", "_").replace("(", "").replace(")", "").replace("–", "-").replace("/", "-")
                    branch_xlsx = build_branch_by_branch_workbook(
                        per_branch_tables, f"{comp_year} {comp_quarter}")
                    st.download_button(
                        "Download branch-by-branch performance (Excel)",
                        data=branch_xlsx,
                        file_name=f"Playbet_Branch_by_Branch_{safe}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
