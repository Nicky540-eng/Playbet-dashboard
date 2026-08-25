"""
comparison_excel.py
Builds an executive-grade comparison workbook for the Playbet quarterly comparison.

Structure:
  • "Overview" cover sheet — branded banner, scope, headline KPIs, and the key findings
    (busiest/least-busy month, best/worst value month, biggest riser & faller).
  • One sheet per analysis — branded banner, subtitle, a formatted table with a bold
    TOTAL row where meaningful, zebra striping, conditional colour on change/rank columns.

No charts (removed by request). Professional Arial throughout. Returns bytes.
"""

from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter

MONTH_NAMES = ["January", "February", "March", "April", "May", "June",
               "July", "August", "September", "October", "November", "December"]

# --- palette (Playbet slate + emerald, matching the dashboard) ---
NAVY      = "0F172A"   # slate-900 banner
SLATE     = "1F3864"
EMERALD   = "10B981"
EMERALD_D = "059669"
GREY_HEAD = "334155"   # table header
ZEBRA     = "F1F5F9"   # light row stripe
POS_GREEN = "16A34A"
NEG_RED   = "C0392B"
AMBER     = "F59E0B"
TEXT_DARK = "0F172A"
TEXT_MUTE = "64748B"

WHITE = "FFFFFF"
THIN  = Side(style="thin", color="D9E2EC")
MED   = Side(style="medium", color=SLATE)

F_TITLE   = Font(name="Arial", size=20, bold=True, color=WHITE)
F_SUBTLE  = Font(name="Arial", size=10, color="CBD5E1")
F_BANNER  = Font(name="Arial", size=14, bold=True, color=WHITE)
F_SUBHEAD = Font(name="Arial", size=10, italic=True, color=TEXT_MUTE)
F_HEAD    = Font(name="Arial", size=10, bold=True, color=WHITE)
F_BASE    = Font(name="Arial", size=10, color=TEXT_DARK)
F_BOLD    = Font(name="Arial", size=10, bold=True, color=TEXT_DARK)
F_KPI_LBL = Font(name="Arial", size=9,  bold=True, color=TEXT_MUTE)
F_KPI_VAL = Font(name="Arial", size=18, bold=True, color=SLATE)
F_FIND    = Font(name="Arial", size=11, bold=True, color=TEXT_DARK)
F_FIND_L  = Font(name="Arial", size=9,  bold=True, color=EMERALD_D)

FILL_NAVY    = PatternFill("solid", fgColor=NAVY)
FILL_BANNER  = PatternFill("solid", fgColor=SLATE)
FILL_HEAD    = PatternFill("solid", fgColor=GREY_HEAD)
FILL_ZEBRA   = PatternFill("solid", fgColor=ZEBRA)
FILL_EMERALD = PatternFill("solid", fgColor=EMERALD)
FILL_TOTAL   = PatternFill("solid", fgColor="E2E8F0")
FILL_KPI     = PatternFill("solid", fgColor="F8FAFC")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left", vertical="center")
RIGHT  = Alignment(horizontal="right", vertical="center")


def _is_pct(colname):
    c = str(colname)
    return "%" in c or "GWM" in c


def _is_money(colname):
    c = str(colname).lower()
    return any(k in c for k in ("paid in", "gross win", "paid out", "net win", "value"))


def _banner(ws, ncols, title, subtitle):
    """Slate banner across the table width + italic subtitle line."""
    last = get_column_letter(max(ncols, 3))
    ws.merge_cells(f"A1:{last}1")
    ws["A1"] = title
    ws["A1"].font = F_BANNER
    ws["A1"].fill = FILL_BANNER
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30
    ws.merge_cells(f"A2:{last}2")
    ws["A2"] = subtitle
    ws["A2"].font = F_SUBHEAD
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 18


def _write_table(ws, df, start_row, total_row=True, rank_cols=None, change_cols=None,
                 month_cols=None):
    """Write a styled table. Returns the row after the table.
    If month_cols is given, in each row the highest value among those columns is
    filled green and the lowest red."""
    rank_cols = rank_cols or []
    change_cols = change_cols or []
    month_cols = month_cols or []
    ncols = len(df.columns)
    GREEN_FILL = PatternFill("solid", fgColor="16A34A")
    RED_FILL = PatternFill("solid", fgColor="C0392B")
    WHITE_BOLD = Font(name="Arial", size=10, bold=True, color="FFFFFF")

    present_months = [c for c in month_cols if c in df.columns]
    row_hilo = []
    if present_months:
        for _, row in df.iterrows():
            vals = pd.to_numeric(pd.Series({c: row[c] for c in present_months}), errors="coerce")
            if vals.notna().any() and vals.max() != vals.min():
                row_hilo.append((vals.idxmax(), vals.idxmin()))
            else:
                row_hilo.append((None, None))

    # header
    hr = start_row
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=hr, column=j, value=str(col))
        c.font = F_HEAD
        c.fill = FILL_HEAD
        c.alignment = CENTER
        c.border = Border(left=THIN, right=THIN, top=MED, bottom=MED)
    ws.row_dimensions[hr].height = 26

    # body
    r = hr + 1
    for i, (_, row) in enumerate(df.iterrows()):
        zebra = (i % 2 == 1)
        for j, (col, val) in enumerate(zip(df.columns, row), start=1):
            v = val.item() if hasattr(val, "item") else val
            if pd.isna(v):
                v = ""
            cell = ws.cell(row=r, column=j, value=v)
            cell.font = F_BASE
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            if zebra:
                cell.fill = FILL_ZEBRA
            # alignment + number format
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                cell.alignment = RIGHT
                if _is_pct(col):
                    cell.number_format = '0.00"%"'
                elif _is_money(col):
                    cell.number_format = 'R #,##0.00'
                else:
                    cell.number_format = "#,##0"
            else:
                cell.alignment = LEFT
            # conditional colour on change columns
            if col in change_cols and isinstance(v, (int, float)):
                if v > 0:
                    cell.font = Font(name="Arial", size=10, bold=True, color=POS_GREEN)
                elif v < 0:
                    cell.font = Font(name="Arial", size=10, bold=True, color=NEG_RED)
            # colour the rank/flag text columns
            if col in rank_cols and isinstance(v, str) and v:
                low = v.lower()
                if "busiest" in low or "best" in low or "increase" in low:
                    cell.font = Font(name="Arial", size=10, bold=True, color=EMERALD_D)
                elif "least" in low or "worst" in low or "decrease" in low:
                    cell.font = Font(name="Arial", size=10, bold=True, color=NEG_RED)
            # highest month figure green, lowest red (per row)
            if row_hilo:
                hi, lo = row_hilo[i]
                if col == hi:
                    cell.fill = GREEN_FILL
                    cell.font = WHITE_BOLD
                elif col == lo:
                    cell.fill = RED_FILL
                    cell.font = WHITE_BOLD
        r += 1

    # TOTAL row (numeric columns only, first column labelled TOTAL)
    if total_row and len(df) > 0:
        numeric_cols = [c for c in df.columns
                        if pd.api.types.is_numeric_dtype(df[c]) and not _is_pct(c)]
        if numeric_cols:
            for j, col in enumerate(df.columns, start=1):
                cell = ws.cell(row=r, column=j)
                cell.fill = FILL_TOTAL
                cell.border = Border(left=THIN, right=THIN, top=MED, bottom=MED)
                if j == 1:
                    cell.value = "TOTAL"
                    cell.font = F_BOLD
                    cell.alignment = LEFT
                elif col in numeric_cols:
                    col_letter = get_column_letter(j)
                    cell.value = f"=SUM({col_letter}{hr+1}:{col_letter}{r-1})"
                    cell.font = F_BOLD
                    cell.alignment = RIGHT
                    cell.number_format = 'R #,##0.00' if _is_money(col) else "#,##0"
            r += 1

    # widths
    for j, col in enumerate(df.columns, start=1):
        vals = [str(col)] + [str(v) for v in df.iloc[:, j - 1].tolist()]
        width = min(max(len(x) for x in vals) + 3, 40)
        ws.column_dimensions[get_column_letter(j)].width = max(width, 11)

    ws.freeze_panes = ws.cell(row=hr + 1, column=1)
    ws.sheet_view.showGridLines = False
    return r + 1


def _sheet(wb, tab, df, title, subtitle, total_row=True, rank_cols=None, change_cols=None):
    ws = wb.create_sheet(title=tab[:31])
    _banner(ws, len(df.columns), title, subtitle)
    _write_table(ws, df, start_row=4, total_row=total_row,
                 rank_cols=rank_cols, change_cols=change_cols)
    return ws


def _overview(wb, meta, kpis, findings):
    ws = wb.create_sheet(title="Overview")
    ws.sheet_view.showGridLines = False
    # brand banner
    ws.merge_cells("A1:F2")
    ws["A1"] = "PLAYBET"
    ws["A1"].font = Font(name="Arial", size=26, bold=True, color=WHITE)
    ws["A1"].fill = FILL_NAVY
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for r in (1, 2):
        for c in range(1, 7):
            ws.cell(row=r, column=c).fill = FILL_NAVY
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 20
    ws.merge_cells("A3:F3")
    ws["A3"] = "Quarterly Performance Comparison"
    ws["A3"].font = Font(name="Arial", size=15, bold=True, color=SLATE)
    ws.merge_cells("A4:F4")
    ws["A4"] = meta
    ws["A4"].font = F_SUBHEAD

    # KPI cards row
    row = 6
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "HEADLINE FIGURES"
    ws[f"A{row}"].font = F_KPI_LBL
    row += 1
    col = 1
    for label, value in kpis.items():
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        lc = ws.cell(row=row, column=col, value=label)
        lc.font = F_KPI_LBL
        lc.fill = FILL_KPI
        lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
        vc = ws.cell(row=row + 1, column=col, value=value)
        vc.font = F_KPI_VAL
        vc.fill = FILL_KPI
        vc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        for rr in (row, row + 1):
            for cc in (col, col + 1):
                ws.cell(row=rr, column=cc).border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        col += 2
        if col > 5:
            col = 1
            row += 3
    row += 3

    # Key findings
    ws.merge_cells(f"A{row}:F{row}")
    fh = ws.cell(row=row, column=1, value="KEY FINDINGS")
    fh.font = F_BANNER
    fh.fill = FILL_BANNER
    fh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 26
    row += 2
    for label, text in findings:
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = F_FIND_L
        lc.alignment = Alignment(horizontal="left", vertical="top", indent=1)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        tc = ws.cell(row=row, column=2, value=text)
        tc.font = F_BASE
        tc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[row].height = 30
        row += 1

    ws.column_dimensions["A"].width = 22
    for c in "BCDEF":
        ws.column_dimensions[c].width = 20
    return ws


def _fmt_money(x):
    try:
        return f"R {x:,.0f}"
    except Exception:
        return str(x)


def build_comparison_workbook(tables: dict, meta: str) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    bw   = tables.get("Best-Worst Month", pd.DataFrame())
    gid  = tables.get("Games Increase-Decrease", pd.DataFrame())
    gbm  = tables.get("Games Betslips by Month", pd.DataFrame())
    cbm  = tables.get("Cashier Betslips by Month", pd.DataFrame())

    # ---- derive headline KPIs + findings for the Overview ----
    kpis = {}
    findings = []
    if not bw.empty:
        if "Betslips" in bw.columns:
            kpis["Total Betslips"] = int(bw["Betslips"].sum())
        if "Paid In" in bw.columns:
            kpis["Total Paid In"] = _fmt_money(bw["Paid In"].sum())
        if "Gross Win" in bw.columns:
            kpis["Total Gross Win"] = _fmt_money(bw["Gross Win"].sum())
        if "Volume" in bw.columns:
            busiest = bw.loc[bw["Volume"] == "Busiest", "Month"]
            quiet   = bw.loc[bw["Volume"] == "Least busy", "Month"]
            if len(busiest):
                findings.append(("BUSIEST", f"{busiest.iloc[0]} had the highest betslip volume in the quarter."))
            if len(quiet):
                findings.append(("QUIETEST", f"{quiet.iloc[0]} had the lowest betslip volume in the quarter."))
        if "Value" in bw.columns:
            best  = bw.loc[bw["Value"] == "Best value", "Month"]
            worst = bw.loc[bw["Value"] == "Worst value", "Month"]
            if len(best):
                findings.append(("BEST VALUE", f"{best.iloc[0]} produced the most gross win (best value month)."))
            if len(worst):
                findings.append(("WORST VALUE", f"{worst.iloc[0]} produced the least gross win (worst value month)."))
    if not gid.empty and "Change in Betslips" in gid.columns:
        top = gid.iloc[0]
        bottom = gid.iloc[-1]
        findings.append(("TOP RISER", f"{top['Game']} rose {int(top['Change in Betslips']):+,} betslips ({top['Change (%)']:+.1f}%) from first to last month."))
        if bottom["Change in Betslips"] < 0:
            findings.append(("TOP FALLER", f"{bottom['Game']} fell {int(bottom['Change in Betslips']):+,} betslips ({bottom['Change (%)']:+.1f}%) over the quarter."))
    if not gbm.empty:
        kpis.setdefault("Games Tracked", len(gbm))
    if not cbm.empty:
        kpis.setdefault("Cashiers Tracked", len(cbm))

    _overview(wb, meta, kpis, findings)

    # ---- analysis sheets ----
    SPECS = [
        ("Best-Worst Month", "Best-Worst Month",
         "Busiest & Least-Busy Month · Best & Worst Value Month",
         "Busiest/least-busy ranked by betslip volume; best/worst by gross win. " + meta,
         False, ["Volume", "Value"], []),
        ("Games Betslips by Month", "Games Betslips by Month",
         "Games — Betslip Count by Month",
         "Each game's betslips side-by-side across the quarter's three months. " + meta,
         True, [], []),
        ("Games Increase-Decrease", "Games Increase-Decrease",
         "Games — Betslip Increase / Decrease",
         "First month vs last month; green = growth, red = decline. " + meta,
         False, ["Trend"], ["Change in Betslips", "Change (%)"]),
        ("Games per Branch", "Games per Branch",
         "Games per Branch",
         "Betslips, Paid In, Gross Win and Gross Win Margin % per game, per branch. " + meta,
         False, [], []),
        ("Games GWM", "Games GWM",
         "Games — Gross Win Margin %",
         "Gross Win Margin % = Gross Win ÷ Paid In, by month and for the whole quarter. " + meta,
         False, [], []),
        ("Cashier Betslips by Month", "Cashier Betslips by Month",
         "Cashiers — Betslip Count by Month",
         "Each cashier's betslips side-by-side across the quarter. " + meta,
         True, [], []),
        ("Cashier GWM", "Cashier GWM",
         "Cashiers — Gross Win Margin %",
         "Per-month GW Margin %; quarter column is paid-in weighted. " + meta,
         False, [], []),
    ]
    for key, tab, title, subtitle, total_row, rank_cols, change_cols in SPECS:
        df = tables.get(key)
        if df is None or df.empty:
            continue
        _sheet(wb, tab, df, title, subtitle,
               total_row=total_row, rank_cols=rank_cols, change_cols=change_cols)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_branch_by_branch_workbook(per_branch_tables: dict, meta: str,
                                    cover=None) -> bytes:
    """Branch-by-branch workbook: a rich cover page plus one sheet per branch.

    per_branch_tables: {branch_name: {table_name: DataFrame, ...}, ...}
    cover (optional): {
        'branch_summary': DataFrame with columns
             Branch, Betslips, Paid In, Paid Out, Gross Win Margin %, Net Win Margin %
        'games_per_branch': DataFrame with columns
             Branch, Game, Paid In, Paid Out
    }
    """
    wb = Workbook()
    wb.remove(wb.active)

    ov = wb.create_sheet(title="Cover")
    ov.sheet_view.showGridLines = False
    ov.merge_cells("A1:H2")
    ov["A1"] = "PLAYBET"
    ov["A1"].font = Font(name="Arial", size=26, bold=True, color=WHITE)
    ov["A1"].fill = FILL_NAVY
    ov["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for rr in (1, 2):
        for c in range(1, 9):
            ov.cell(row=rr, column=c).fill = FILL_NAVY
    ov.row_dimensions[1].height = 26
    ov.merge_cells("A3:H3")
    ov["A3"] = "Branch Performance"
    ov["A3"].font = Font(name="Arial", size=15, bold=True, color=SLATE)
    ov.merge_cells("A4:H4")
    ov["A4"] = meta
    ov["A4"].font = F_SUBHEAD
    for w, col in zip([20, 16, 16, 16, 16, 16, 16, 16], "ABCDEFGH"):
        ov.column_dimensions[col].width = w

    row = 6
    if cover and cover.get("branch_summary") is not None and not cover["branch_summary"].empty:
        bs = cover["branch_summary"].set_index("Branch")

        # Fixed pairings requested: Malvern vs Randburg, Potchefstroom vs Pretoria,
        # White River on its own.
        pairings = [("Malvern", "Randburg"),
                    ("Potchefstroom", "Pretoria"),
                    ("White River", None)]
        metrics = ["Betslips", "Paid In", "Paid Out",
                   "Gross Win Margin %", "Net Win Margin %"]

        def _fmt(metric, val):
            if val is None or (isinstance(val, float) and pd.isna(val)):
                return "—"
            if "%" in metric:
                return f"{val:,.2f}%"
            if metric in ("Paid In", "Paid Out"):
                return f"R {val:,.0f}"
            return f"{val:,.0f}"

        for left, right in pairings:
            have_left = left in bs.index
            have_right = right in bs.index if right else False
            if not have_left and not have_right:
                continue
            title = f"{left}  vs  {right}" if right else f"{left}"
            ov.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            tcell = ov.cell(row=row, column=1, value=title)
            tcell.font = Font(name="Arial", size=12, bold=True, color=WHITE)
            for c in range(1, 5):
                ov.cell(row=row, column=c).fill = FILL_NAVY
            row += 1

            # header
            ov.cell(row=row, column=1, value="Metric").font = F_BOLD
            ov.cell(row=row, column=2, value=left).font = F_BOLD
            if right:
                ov.cell(row=row, column=3, value=right).font = F_BOLD
                ov.cell(row=row, column=4, value="Leader").font = F_BOLD
            for c in range(1, 5 if right else 3):
                ov.cell(row=row, column=c).fill = FILL_HEAD
                ov.cell(row=row, column=c).font = Font(name="Arial", size=10, bold=True, color=WHITE)
            row += 1

            for m in metrics:
                lv = bs.loc[left, m] if have_left and m in bs.columns else None
                ov.cell(row=row, column=1, value=m).font = F_BASE
                ov.cell(row=row, column=2, value=_fmt(m, lv)).font = F_BASE
                if right:
                    rv = bs.loc[right, m] if have_right and m in bs.columns else None
                    ov.cell(row=row, column=3, value=_fmt(m, rv)).font = F_BASE
                    # leader (higher is better for all these metrics)
                    leader = ""
                    if lv is not None and rv is not None and not pd.isna(lv) and not pd.isna(rv):
                        if lv > rv:
                            leader = left
                            ov.cell(row=row, column=2).fill = PatternFill("solid", fgColor="16A34A")
                            ov.cell(row=row, column=2).font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                        elif rv > lv:
                            leader = right
                            ov.cell(row=row, column=3).fill = PatternFill("solid", fgColor="16A34A")
                            ov.cell(row=row, column=3).font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                    ov.cell(row=row, column=4, value=leader).font = F_BASE
                for c in range(1, 5 if right else 3):
                    ov.cell(row=row, column=c).border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
                row += 1
            row += 1  # gap between pairings

    # Game-by-game per branch: PAIRED side-by-side comparison (Option C).
    # Malvern vs Randburg, Potchefstroom vs Pretoria, White River on its own.
    # Each game is a row; the two branches' Paid In / Paid Out sit side by side.
    if cover and cover.get("games_per_branch") is not None and not cover["games_per_branch"].empty:
        gpb = cover["games_per_branch"].copy()
        gpb["Paid Out"] = gpb["Paid Out"].abs()
        present = set(gpb["Branch"])
        row += 1
        ov.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        h = ov.cell(row=row, column=1, value="Game-by-Game Comparison — Paid In vs Paid Out")
        h.font = Font(name="Arial", size=13, bold=True, color=SLATE)
        row += 2

        pairings = [("Malvern", "Randburg"),
                    ("Potchefstroom", "Pretoria"),
                    ("White River", None)]

        def _games_map(branch):
            sub = gpb[gpb["Branch"] == branch]
            return {r["Game"]: (float(r["Paid In"]), float(r["Paid Out"])) for _, r in sub.iterrows()}

        for left, right in pairings:
            if left not in present and (right is None or right not in present):
                continue
            lmap = _games_map(left)
            rmap = _games_map(right) if right else {}
            games = sorted(set(lmap) | set(rmap),
                           key=lambda g: lmap.get(g, (0, 0))[0], reverse=True)
            if not games:
                continue

            ncol = 5 if right else 3
            # Title band
            ov.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncol)
            title = f"{left}  vs  {right}" if right else left
            tc = ov.cell(row=row, column=1, value=title)
            tc.font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
            for c in range(1, ncol + 1):
                ov.cell(row=row, column=c).fill = FILL_NAVY
            row += 1

            # Group header row: branch names spanning their two columns
            ov.cell(row=row, column=1, value="").fill = FILL_HEAD
            ov.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
            gc = ov.cell(row=row, column=2, value=left)
            gc.fill = FILL_HEAD; gc.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            gc.alignment = Alignment(horizontal="center")
            ov.cell(row=row, column=3).fill = FILL_HEAD
            if right:
                ov.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
                gc2 = ov.cell(row=row, column=4, value=right)
                gc2.fill = FILL_HEAD; gc2.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                gc2.alignment = Alignment(horizontal="center")
                ov.cell(row=row, column=5).fill = FILL_HEAD
            ov.cell(row=row, column=1).fill = FILL_HEAD
            row += 1

            # Column header row
            heads = ["Game", "Paid In", "Paid Out"] + (["Paid In", "Paid Out"] if right else [])
            for j, lab in enumerate(heads, start=1):
                cell = ov.cell(row=row, column=j, value=lab)
                cell.fill = FILL_HEAD
                cell.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
            row += 1

            for g in games:
                lpi, lpo = lmap.get(g, (0.0, 0.0))
                ov.cell(row=row, column=1, value=g).font = F_BASE
                ov.cell(row=row, column=2, value=round(lpi, 0)).font = F_BASE
                ov.cell(row=row, column=3, value=round(lpo, 0)).font = F_BASE
                if right:
                    rpi, rpo = rmap.get(g, (0.0, 0.0))
                    ov.cell(row=row, column=4, value=round(rpi, 0)).font = F_BASE
                    ov.cell(row=row, column=5, value=round(rpo, 0)).font = F_BASE
                    # highlight the branch with higher Paid In for this game
                    if lpi != rpi:
                        winner_col = 2 if lpi > rpi else 4
                        ov.cell(row=row, column=winner_col).fill = PatternFill("solid", fgColor="16A34A")
                        ov.cell(row=row, column=winner_col).font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                for c in range(2, ncol + 1):
                    ov.cell(row=row, column=c).number_format = 'R #,##0'
                for c in range(1, ncol + 1):
                    ov.cell(row=row, column=c).border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
                row += 1
            row += 1

    # Sheet name suffixes per table (kept short so branch+suffix fits 31 chars).
    order = [
        ("Games Betslips by Month", "Games"),
        ("Games GWM", "Games GWM%"),
        ("Cashier Betslips by Month", "Cashiers"),
        ("Cashier GWM", "Cashiers GWM%"),
    ]
    titles = {
        "Games Betslips by Month": "Games — Betslip Count by Month",
        "Games GWM": "Games — Gross Win Margin %",
        "Cashier Betslips by Month": "Cashiers — Betslip Count by Month",
        "Cashier GWM": "Cashiers — Gross Win Margin %",
    }

    for branch, tables in per_branch_tables.items():
        for key, suffix in order:
            df = tables.get(key)
            if df is None or df.empty:
                continue
            # Sheet name like "Malvern · Games" trimmed to 31 chars.
            raw = f"{branch} {suffix}"
            tab = raw[:31]
            base = tab
            n = 2
            while tab in wb.sheetnames:
                tab = f"{base[:28]}_{n}"
                n += 1
            ws = wb.create_sheet(title=tab)
            _banner(ws, len(df.columns), f"{branch} — {titles.get(key, key)}", meta)
            # Columns to highlight highest-green/lowest-red per row:
            #  - Betslip tables: the bare month columns (January, February, ...)
            #  - GWM% tables: the monthly "... % January" columns, but NOT the Quarter column
            if "Betslip" in key:
                mcols = [c for c in MONTH_NAMES if c in df.columns]
            elif "GWM" in key:
                mcols = [c for c in df.columns
                         if any(m in str(c) for m in MONTH_NAMES) and "Quarter" not in str(c)]
            else:
                mcols = []
            _write_table(ws, df, start_row=4, total_row=("Betslip" in key),
                         month_cols=mcols)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
